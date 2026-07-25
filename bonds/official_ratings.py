#!/usr/bin/env python3
"""Current Russian bond ratings from official agency publications.

The adapter deliberately matches debt issues by ISIN.  It never infers an issue
rating from an issuer name.  For an issue rated by several agencies the selected
rating is the lowest current rating; every source record is retained for the UI.
"""
from __future__ import annotations

import html
import json
import os
import re
import subprocess
import tempfile
import time
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Callable, Iterable

import requests

UA = (
    "dividend-factor-strategies/1.0 "
    "(+https://github.com/eremkindv91/dividend-factor-strategies)"
)
ACRA_UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/126 Safari/537.36"
TIMEOUT = 45
RATE_LIMIT_SECONDS = 0.15

ACRA_API = "https://www.acra-ratings.ru/local/ajax/get_rate_emission.php"
ACRA_BASE = "https://www.acra-ratings.ru"
EXPERT_BASE = "https://raexpert.ru"
NKR_BASE = "https://ratings.ru"
NKR_XLSX = f"{NKR_BASE}/issues.php"
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_CACHE = os.path.join(REPO, "data", "bond_finder", "official_ratings_cache.json")

RATING_RANK = {
    "AAA": 20,
    "AA+": 19,
    "AA": 18,
    "AA-": 17,
    "A+": 16,
    "A": 15,
    "A-": 14,
    "BBB+": 13,
    "BBB": 12,
    "BBB-": 11,
    "BB+": 10,
    "BB": 9,
    "BB-": 8,
    "B+": 7,
    "B": 6,
    "B-": 5,
    "CCC+": 4,
    "CCC": 3,
    "CCC-": 2,
    "CC": 1,
    "C": 0,
}
ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{10}$")
RATING_RE = re.compile(
    r"^(AAA|AA[+-]?|A[+-]?|BBB[+-]?|BB[+-]?|B[+-]?|CCC[+-]?|CC|C)$"
)


def normalize_rating(value: object) -> str | None:
    """Convert ACRA/Expert RA/NKR national scales to one canonical scale."""
    if value is None:
        return None
    raw = str(value).strip().upper()
    if not raw or any(x in raw for x in ("ОТОЗВАН", "WITHDRAW", "ПОГАШЕН", "НЕТ РЕЙТИНГА")):
        return None
    raw = raw.translate(str.maketrans({"А": "A", "В": "B", "С": "C"}))
    raw = re.sub(r"\s+", "", raw)
    raw = re.sub(r"\((RU|РУ)\)$", "", raw)
    raw = re.sub(r"\.(RU|РУ)$", "", raw)
    raw = re.sub(r"\|(RU|РУ)\|$", "", raw)
    if raw.startswith("RU") and RATING_RE.fullmatch(raw[2:]):
        raw = raw[2:]
    return raw if raw in RATING_RANK else None


def _iso_date(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _record(
    *,
    isin: object,
    raw_rating: object,
    agency: str,
    rating_date: object,
    source_url: str,
    press_release_url: str | None = None,
) -> dict | None:
    sid = str(isin or "").strip().upper()
    rating = normalize_rating(raw_rating)
    if not ISIN_RE.fullmatch(sid) or rating is None:
        return None
    return {
        "isin": sid,
        "rating": rating,
        "raw_rating": str(raw_rating).strip(),
        "rating_agency": agency,
        "rating_scope": "issue",
        "rating_date": _iso_date(rating_date),
        "rating_source_url": source_url,
        "rating_press_release_url": press_release_url,
    }


# ACRA присылает ТОЛЬКО листовой сертификат, без промежуточного CA (ошибка их конфигурации:
# openssl verify → 21 "unable to verify the first certificate"). Штатное решение — AIA chasing:
# берём URI издателя из расширения Authority Information Access самого сертификата, скачиваем
# промежуточный (публичный GlobalSign RSA OV SSL CA 2018, цепляется к корню из доверенного
# хранилища) и добавляем его к CA-бандлу. Проверка TLS при этом ОСТАЁТСЯ ВКЛЮЧЁННОЙ —
# отключать её нельзя, это дало бы MITM-дыру ради удобства.
_CA_BUNDLE_CACHE: str | None = None


def _acra_ca_bundle() -> str | None:
    """certifi-бандл + промежуточный CA из AIA сертификата ACRA. None, если собрать не удалось."""
    global _CA_BUNDLE_CACHE
    if _CA_BUNDLE_CACHE is not None:
        return _CA_BUNDLE_CACHE or None
    try:
        import ssl
        import certifi
        # 1) листовой сертификат → URI издателя (AIA)
        leaf = ssl.get_server_certificate(("www.acra-ratings.ru", 443))
        text = subprocess.run(["openssl", "x509", "-noout", "-text"], input=leaf,
                              capture_output=True, text=True, timeout=20, check=False).stdout
        m = re.search(r"CA Issuers - URI:(\S+)", text)
        if not m:
            _CA_BUNDLE_CACHE = ""
            return None
        # 2) промежуточный CA (DER или PEM) → PEM
        der = subprocess.run(["curl", "--fail", "--silent", "--max-time", "20", m.group(1)],
                             capture_output=True, timeout=25, check=False).stdout
        if not der:
            _CA_BUNDLE_CACHE = ""
            return None
        pem = subprocess.run(["openssl", "x509", "-inform", "DER"], input=der,
                             capture_output=True, timeout=20, check=False).stdout.decode("utf-8", "replace")
        if "BEGIN CERTIFICATE" not in pem:
            pem = der.decode("utf-8", "replace")
        if "BEGIN CERTIFICATE" not in pem:
            _CA_BUNDLE_CACHE = ""
            return None
        # 3) бандл = доверенные корни + этот промежуточный
        fd, path = tempfile.mkstemp(prefix="acra-ca-", suffix=".pem")
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(Path(certifi.where()).read_text(encoding="utf-8"))
            handle.write("\n")
            handle.write(pem)
        _CA_BUNDLE_CACHE = path
        return path
    except Exception:  # noqa: BLE001  (сеть/openssl недоступны — просто нет бандла)
        _CA_BUNDLE_CACHE = ""
        return None


def _curl_post_json(url: str, payload: dict, timeout: int) -> dict:
    """TLS-verified curl fallback for ACRA's incomplete cert chain (AIA-догруженный бандл)."""
    bundle = _acra_ca_bundle()
    cmd = [
        "curl",
        "--fail",
        "--silent",
        "--show-error",
        "--max-time",
        str(timeout),
        *(["--cacert", bundle] if bundle else []),
        "-A",
        ACRA_UA,
        "-H",
        "Content-Type: application/json",
        "--data-binary",
        "@-",
        url,
    ]
    proc = subprocess.run(
        cmd,
        input=json.dumps(payload),
        capture_output=True,
        text=True,
        timeout=timeout + 5,
        check=False,
    )
    if proc.returncode != 0:
        raise RuntimeError(f"curl HTTP error: {proc.stderr.strip()[:180]}")
    return json.loads(proc.stdout)


def _post_json(session: requests.Session, url: str, payload: dict, timeout: int = TIMEOUT) -> dict:
    try:
        response = session.post(url, json=payload, timeout=timeout)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.SSLError:
        # сначала пробуем тот же запрос с AIA-догруженным промежуточным CA (верификация включена)
        bundle = _acra_ca_bundle()
        if bundle:
            try:
                response = session.post(url, json=payload, timeout=timeout, verify=bundle)
                response.raise_for_status()
                return response.json()
            except requests.exceptions.RequestException:
                pass
        # No certificate checks are disabled. curl uses the host trust store and
        # succeeds when Python/OpenSSL cannot build ACRA's intermediate chain.
        return _curl_post_json(url, payload, timeout)


def parse_acra_payload(payload: dict) -> list[dict]:
    rows = []
    for item in (payload.get("data") or {}).get("items") or []:
        info = item.get("info") or {}
        emission = (info.get("emission") or {}).get("value") or {}
        rate = (info.get("rate") or {}).get("value") or {}
        release = (info.get("pressRelease") or {}).get("value") or {}
        if rate.get("is_withdrawn"):
            continue
        timestamp = release.get("timestamp")
        rating_date = None
        if isinstance(timestamp, (int, float)):
            rating_date = datetime.fromtimestamp(timestamp, tz=timezone.utc).date()
        row = _record(
            isin=emission.get("isin"),
            raw_rating=rate.get("title"),
            agency="АКРА",
            rating_date=rating_date,
            source_url=ACRA_BASE + str(emission.get("url") or ""),
            press_release_url=(ACRA_BASE + str(release.get("url"))) if release.get("url") else None,
        )
        if row:
            rows.append(row)
    return rows


def load_acra_ratings(session: requests.Session | None = None) -> list[dict]:
    session = session or requests.Session()
    session.headers.update({"User-Agent": ACRA_UA, "Accept": "application/json"})
    rows: list[dict] = []
    page = 1
    pages = 1
    while page <= pages:
        request = {
            "text": "",
            "sectors": [],
            "activities": [],
            "countries": [],
            "forecasts": [],
            "on_revision": 0,
            "rating_scale": 0,
            "rate_from": 0,
            "rate_to": 0,
            "vexel_types": "",
            "issue_type": "",
            "debt_types": "",
            "page": page,
            "sort": "",
            "count": 1000,
        }
        payload = _post_json(session, ACRA_API, request)
        if not payload.get("success"):
            raise RuntimeError("ACRA API returned success=false")
        data = payload.get("data") or {}
        pages = max(1, min(int(data.get("count_pages") or 1), 10))
        rows.extend(parse_acra_payload(payload))
        page += 1
        if page <= pages:
            time.sleep(RATE_LIMIT_SECONDS)
    if not rows:
        raise RuntimeError("ACRA issue list is empty")
    return rows


def parse_nkr_workbook(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - guarded by workflow dependency
        raise RuntimeError("openpyxl is required for the official NKR workbook") from exc
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(x or "").strip() for x in next(iterator)]
    rows = []
    for values in iterator:
        item = dict(zip(headers, values))
        raw = item.get("Rating")
        issue_id = item.get("ID")
        try:
            issue_id = str(int(issue_id))
        except (TypeError, ValueError):
            issue_id = ""
        press = str(item.get("Press release") or "").strip()
        if press and not press.startswith(("http://", "https://")):
            press = "https://" + press.lstrip("/")
        row = _record(
            isin=item.get("ISIN"),
            raw_rating=raw,
            agency="НКР",
            rating_date=item.get("Rating Date"),
            source_url=f"{NKR_BASE}/ratings/issues/{issue_id}/" if issue_id else press,
            press_release_url=press or None,
        )
        if row:
            rows.append(row)
    return rows


def load_nkr_ratings(session: requests.Session | None = None) -> list[dict]:
    session = session or requests.Session()
    session.headers.update({"User-Agent": UA})
    response = session.get(NKR_XLSX, timeout=TIMEOUT)
    response.raise_for_status()
    if not response.content.startswith(b"PK"):
        raise RuntimeError("NKR issues.php did not return an XLSX workbook")
    rows = parse_nkr_workbook(response.content)
    if not rows:
        raise RuntimeError("NKR issue workbook has no active ratings")
    return rows


def _clean_html(value: str) -> str:
    return html.unescape(re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", value))).strip()


def parse_expert_rating_page(content: str) -> list[dict]:
    rows = []
    for tr in re.findall(r"<tr(?:\s[^>]*)?>(.*?)</tr>", content, re.I | re.S):
        path_match = re.search(r'href="(/database/securities/bonds/\d+)/?"', tr, re.I)
        cells = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.I | re.S)
        if not path_match or len(cells) < 4:
            continue
        raw_rating = _clean_html(cells[1])
        if normalize_rating(raw_rating) is None:
            continue
        release_match = re.search(r'href="(/releases/[^"]+)"', cells[3], re.I)
        rows.append(
            {
                "detail_path": path_match.group(1).rstrip("/"),
                "raw_rating": raw_rating,
                "rating_date": _iso_date(_clean_html(cells[3])),
                "press_release_url": (
                    EXPERT_BASE + release_match.group(1) if release_match else None
                ),
            }
        )
    return rows


def parse_expert_isin_map(content: str) -> dict[str, str]:
    pairs = re.findall(
        r'href="(/database/securities/bonds/\d+)/?"[^>]*>\s*([A-Z]{2}[A-Z0-9]{10})\s*</a>',
        content,
        re.I | re.S,
    )
    return {path.rstrip("/"): isin.upper() for path, isin in pairs}


def parse_expert_detail_rating(content: str, isin: str, detail_path: str) -> dict | None:
    """Parse the current rating shown on one exact Expert RA issue card."""
    description = re.search(
        r'<meta\s+name="description"\s+content="([^"]+)"', content, re.I
    )
    if not description:
        return None
    raw_match = re.search(
        r"(?<![A-Z0-9])(ru(?:AAA|AA[+-]?|A[+-]?|BBB[+-]?|BB[+-]?|B[+-]?|CCC[+-]?|CC|C))(?![A-Z0-9+-])",
        html.unescape(description.group(1)),
        re.I,
    )
    if not raw_match or normalize_rating(raw_match.group(1)) is None:
        return None

    table = re.search(
        r'<table[^>]*class="[^"]*object-rating-table[^"]*"[^>]*>(.*?)</table>',
        content,
        re.I | re.S,
    )
    if not table:
        return None
    first_row = re.search(r"<tr(?:\s[^>]*)?>(.*?)</tr>", table.group(1), re.I | re.S)
    # Skip the header row when present.
    remaining = table.group(1)[first_row.end() :] if first_row else table.group(1)
    rating_row = re.search(r"<tr(?:\s[^>]*)?>(.*?)</tr>", remaining, re.I | re.S)
    if not rating_row:
        return None
    row_html = rating_row.group(1)
    row_rating = re.search(
        r"(?<![A-Z0-9])(ru(?:AAA|AA[+-]?|A[+-]?|BBB[+-]?|BB[+-]?|B[+-]?|CCC[+-]?|CC|C))(?![A-Z0-9+-])",
        _clean_html(row_html),
        re.I,
    )
    if not row_rating or normalize_rating(row_rating.group(1)) != normalize_rating(raw_match.group(1)):
        return None
    release = re.search(r'href="(/releases/[^"]+)"[^>]*>(\d{2}\.\d{2}\.\d{4})</a>', row_html, re.I)
    return _record(
        isin=isin,
        raw_rating=raw_match.group(1),
        agency="Эксперт РА",
        rating_date=release.group(2) if release else None,
        source_url=EXPERT_BASE + detail_path.rstrip("/") + "/",
        press_release_url=EXPERT_BASE + release.group(1) if release else None,
    )


def load_expert_ra_exact_ratings(
    isins: Iterable[str], session: requests.Session | None = None
) -> list[dict]:
    """Resolve a small set of missing ISINs through exact official issue cards."""
    wanted = {str(x).strip().upper() for x in isins if ISIN_RE.fullmatch(str(x).strip().upper())}
    if not wanted:
        return []
    if len(wanted) > 25:
        raise ValueError("Expert RA exact lookup is limited to 25 ISINs per run")
    session = session or requests.Session()
    session.headers.update({"User-Agent": UA})
    mapping_response = session.get(f"{EXPERT_BASE}/database/securities/bonds/", timeout=TIMEOUT)
    mapping_response.raise_for_status()
    path_by_isin = {isin: path for path, isin in parse_expert_isin_map(mapping_response.text).items()}
    rows = []
    for isin in sorted(wanted):
        detail_path = path_by_isin.get(isin)
        if not detail_path:
            continue
        response = session.get(EXPERT_BASE + detail_path + "/", timeout=TIMEOUT)
        response.raise_for_status()
        row = parse_expert_detail_rating(response.text, isin, detail_path)
        if row:
            rows.append(row)
        time.sleep(RATE_LIMIT_SECONDS)
    return rows


def load_expert_ra_ratings(session: requests.Session | None = None) -> list[dict]:
    session = session or requests.Session()
    session.headers.update({"User-Agent": UA})
    rating_url = f"{EXPERT_BASE}/ratings/debt_inst/"
    response = session.get(rating_url, timeout=TIMEOUT)
    response.raise_for_status()
    page = response.text
    csrf_match = re.search(r"CSRFAjaxTokenPageHash\s*=\s*'([^']+)'", page)
    if not csrf_match:
        raise RuntimeError("Expert RA page token not found")
    hashes = list(dict.fromkeys(re.findall(r"setRatingPageHash\('([^']+)'\)", page)))
    if len(hashes) > 20:
        raise RuntimeError("Expert RA pagination exceeded safety limit")
    raw_rows = parse_expert_rating_page(page)
    for page_hash in hashes:
        set_page = session.post(
            f"{EXPERT_BASE}/ratings/index/ajax-set-rating-page-hash/",
            data={"rating_page_hash": page_hash, "CSRFAjaxToken": csrf_match.group(1)},
            timeout=TIMEOUT,
        )
        set_page.raise_for_status()
        response = session.get(rating_url, timeout=TIMEOUT)
        response.raise_for_status()
        raw_rows.extend(parse_expert_rating_page(response.text))
        time.sleep(RATE_LIMIT_SECONDS)
    if not raw_rows:
        raise RuntimeError("Expert RA current issue rating list is empty")

    mapping_response = session.get(f"{EXPERT_BASE}/database/securities/bonds/", timeout=TIMEOUT)
    mapping_response.raise_for_status()
    isin_by_path = parse_expert_isin_map(mapping_response.text)
    rows = []
    for raw in raw_rows:
        isin = isin_by_path.get(raw["detail_path"])
        if not isin:
            continue
        row = _record(
            isin=isin,
            raw_rating=raw["raw_rating"],
            agency="Эксперт РА",
            rating_date=raw["rating_date"],
            source_url=EXPERT_BASE + raw["detail_path"] + "/",
            press_release_url=raw["press_release_url"],
        )
        if row:
            rows.append(row)
    if not rows:
        raise RuntimeError("Expert RA ratings could not be matched to ISIN")
    return rows


def select_ratings(records: Iterable[dict], checked_at: str | None = None) -> dict[str, dict]:
    checked_at = checked_at or datetime.now(timezone.utc).isoformat(timespec="seconds")
    grouped: dict[str, list[dict]] = {}
    seen = set()
    for row in records:
        key = (row.get("isin"), row.get("rating_agency"), row.get("rating"))
        if key in seen or row.get("rating") not in RATING_RANK:
            continue
        seen.add(key)
        grouped.setdefault(row["isin"], []).append(dict(row))
    selected = {}
    for isin, issue_rows in grouped.items():
        worst_rank = min(RATING_RANK[row["rating"]] for row in issue_rows)
        worst = [row for row in issue_rows if RATING_RANK[row["rating"]] == worst_rank]
        chosen = max(worst, key=lambda row: row.get("rating_date") or "")
        record = dict(chosen)
        record["rating_rank"] = worst_rank
        record["rating_checked_at"] = checked_at
        record["rating_records"] = sorted(
            [
                {
                    key: row.get(key)
                    for key in (
                        "rating",
                        "raw_rating",
                        "rating_agency",
                        "rating_scope",
                        "rating_date",
                        "rating_source_url",
                        "rating_press_release_url",
                        "rating_source_fetched_at",
                    )
                    if row.get(key) is not None
                }
                for row in issue_rows
            ],
            key=lambda row: row.get("rating_agency") or "",
        )
        selected[isin] = record
    return selected


def _read_cache(path: str | None) -> dict:
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as stream:
            payload = json.load(stream)
        return payload if isinstance(payload, dict) else {}
    except (OSError, ValueError):
        return {}


def _cache_age_hours(value: object, now: datetime) -> float | None:
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return max(0.0, (now - parsed.astimezone(timezone.utc)).total_seconds() / 3600)
    except (TypeError, ValueError):
        return None


def _write_cache(path: str | None, source_cache: dict) -> None:
    if not path:
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as stream:
        json.dump({"version": 1, "sources": source_cache}, stream, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, path)


def load_official_ratings(
    isins: Iterable[str] | None = None,
    loaders: dict[str, Callable[[], list[dict]]] | None = None,
    cache_path: str | None = None,
    refresh: bool = True,
    fresh_cache_hours: int = 6,
    max_fallback_age_days: int = 14,
    expert_exact_lookup: bool = True,
) -> tuple[dict[str, dict], dict]:
    """Load official issue ratings; individual source failures are non-fatal."""
    now = datetime.now(timezone.utc)
    checked_at = now.isoformat(timespec="seconds")
    wanted = {str(x).upper() for x in isins or [] if ISIN_RE.fullmatch(str(x).upper())}
    default_loaders = loaders is None
    loaders = loaders or {
        "acra": load_acra_ratings,
        "expert_ra": load_expert_ra_ratings,
        "nkr": load_nkr_ratings,
    }
    cached = _read_cache(cache_path).get("sources") or {}
    statuses = {}
    records = []
    next_cache = dict(cached)
    for source, loader in loaders.items():
        cached_source = cached.get(source) if isinstance(cached, dict) else None
        cached_rows = (cached_source or {}).get("records") or []
        cached_at = (cached_source or {}).get("fetched_at")
        cache_age = _cache_age_hours(cached_at, now)
        if not refresh and cached_rows and cache_age is not None and cache_age <= fresh_cache_hours:
            records.extend(cached_rows)
            statuses[source] = {
                "status": "ok",
                "mode": "fresh_cache",
                "records": len(cached_rows),
                "last_success_at": cached_at,
            }
            continue
        try:
            got = loader()
            for row in got:
                row["rating_source_fetched_at"] = checked_at
            records.extend(got)
            statuses[source] = {
                "status": "ok",
                "mode": "live",
                "records": len(got),
                "last_success_at": checked_at,
            }
            next_cache[source] = {"fetched_at": checked_at, "records": got}
        except Exception as exc:  # noqa: BLE001 - source isolation is intentional
            if cached_rows and cache_age is not None and cache_age <= max_fallback_age_days * 24:
                records.extend(cached_rows)
                statuses[source] = {
                    "status": "cached",
                    "mode": "last_good",
                    "records": len(cached_rows),
                    "last_success_at": cached_at,
                    "error": str(exc)[:180],
                }
            else:
                statuses[source] = {"status": "error", "mode": "unavailable", "records": 0, "error": str(exc)[:180]}
    selected = select_ratings(records, checked_at=checked_at)
    missing = wanted - set(selected)
    if default_loaders and expert_exact_lookup and 0 < len(missing) <= 25:
        try:
            exact = load_expert_ra_exact_ratings(missing)
            for row in exact:
                row["rating_source_fetched_at"] = checked_at
            if exact:
                records.extend(exact)
                expert_cache = next_cache.get("expert_ra") or {}
                combined = list(expert_cache.get("records") or []) + exact
                unique = {
                    (row.get("isin"), row.get("rating_agency"), row.get("rating")): row
                    for row in combined
                }
                next_cache["expert_ra"] = {
                    "fetched_at": expert_cache.get("fetched_at") or checked_at,
                    "records": list(unique.values()),
                }
                statuses.setdefault("expert_ra", {})["exact_records"] = len(exact)
                selected = select_ratings(records, checked_at=checked_at)
        except Exception as exc:  # noqa: BLE001 - bulk agency data remains usable
            statuses.setdefault("expert_ra", {})["exact_error"] = str(exc)[:180]
    _write_cache(cache_path, next_cache)
    if wanted:
        selected = {isin: row for isin, row in selected.items() if isin in wanted}
    meta = {
        "checked_at": checked_at,
        "selection_policy": "lowest_current_official_issue_rating",
        "sources": statuses,
        "sources_ok": sum(1 for row in statuses.values() if row["status"] in ("ok", "cached")),
        "sources_live": sum(1 for row in statuses.values() if row.get("mode") == "live"),
        "records_total": len(records),
        "matched_issues": len(selected),
    }
    return selected, meta


if __name__ == "__main__":
    rating_map, metadata = load_official_ratings(cache_path=DEFAULT_CACHE, refresh=True)
    print(json.dumps({"meta": metadata, "issues": len(rating_map)}, ensure_ascii=False, indent=2))
