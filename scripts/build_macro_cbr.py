#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Макро-слой от Банка России: ключевая ставка (дневная история) + инфляция г/г (месячная).

Зачем частному инвестору именно это:
  • КЛЮЧЕВАЯ СТАВКА — одновременно ставка дисконтирования и альтернатива в виде вклада.
    Пока она 14%, дивидендная доходность 8% проигрывает депозиту, и это главный факт
    при выборе между акциями и вкладом;
  • ИНФЛЯЦИЯ г/г — переводит номинальную доходность в реальную. Дивиденды 10% при
    инфляции 6% — это +4% покупательной способности, а не +10%;
  • РЕАЛЬНАЯ СТАВКА (ставка − инфляция) — насколько жёсткая политика ЦБ сейчас;
  • ЦЕЛЬ ПО ИНФЛЯЦИИ — ориентир, к которому ЦБ ведёт ставку.

Источники (оба официальные, Банк России):
  1. SOAP DailyInfoWebServ, метод KeyRate — дневная история ключевой ставки;
  2. https://www.cbr.ru/hd_base/infl/ — таблица «Инфляция и ключевая ставка» (месяц,
     ставка, инфляция г/г, цель). Инфляции в SOAP-сервисе ЦБ НЕТ (проверено по WSDL:
     есть KeyRate/Ruonia/MKR/Bliquidity и прочие ставки, инфляции среди методов нет),
     поэтому берётся страница статистики. Формат запроса: UniDbQuery.From/To в ДД.ММ.ГГГГ.

Честность:
  • при недоступности источника прошлый файл НЕ перезаписывается (last-good остаётся);
  • данные проходят гейты правдоподобия — мусор не публикуется даже если распарсился;
  • ничего не достраивается и не интерполируется: пропущенный месяц остаётся пропуском.

Запуск:  python scripts/build_macro_cbr.py
Выход:   site/macro_cbr.json
"""
from __future__ import annotations

import json
import os
import re
import tempfile
import subprocess
import ssl
import sys
import time
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta, timezone

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "site", "macro_cbr.json")
MSK = timezone(timedelta(hours=3))

SOAP_URL = "http://www.cbr.ru/DailyInfoWebServ/DailyInfo.asmx"
SOAP_NS = "http://web.cbr.ru/"
INFL_URL = "https://www.cbr.ru/hd_base/infl/"
UA = {"User-Agent": "Mozilla/5.0 (compatible; dividend-factor-strategies/1.0)"}

HISTORY_FROM = "01.01.2015"          # 12 лет — хватает, чтобы увидеть все режимы ДКП
MIN_INFL_ROWS = 60                   # меньше 5 лет — не публикуем
MIN_RATE_ROWS = 30
RATE_RANGE = (0.0, 30.0)             # ключевая ставка вне этого коридора — ошибка парсинга
INFL_RANGE = (-5.0, 30.0)


def fail(msg: str) -> "NoReturn":  # type: ignore[name-defined]
    sys.stderr.write(f"[macro] ОШИБКА: {msg}\n")
    sys.exit(1)


def http(url: str, data: bytes | None = None, headers: dict | None = None,
         retries: int = 4, timeout: int = 45) -> bytes:
    """GET/POST с backoff. ЦБ периодически отвечает медленно — повторяем, а не падаем."""
    last = None
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(url, data=data, headers={**UA, **(headers or {})})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except Exception as e:  # noqa: BLE001
            last = e
            if attempt < retries:
                pause = 2 ** (attempt - 1)
                sys.stderr.write(f"[macro] {url[:60]} не ответил ({e}); повтор через {pause}s\n")
                time.sleep(pause)
    raise RuntimeError(f"{url}: {last}")


# ── Росстат: национальный корень доверия ──────────────────────────────────────
# rosstat.gov.ru выпущен УЦ Минцифры («Russian Trusted Root CA»), которого НЕТ ни в
# certifi, ни в системном хранилище, поэтому обычный запрос падает с
# CERTIFICATE_VERIFY_FAILED. Это не обрыв соединения и не блокировка: сайт доступен,
# просто цепочка не строится. AIA chasing здесь не спасает — недоверен именно КОРЕНЬ.
#
# Решение: собираем ОТДЕЛЬНЫЙ бандл (доверенные корни + корень Минцифры + промежуточный)
# и используем его ТОЛЬКО для запросов к Росстату. Ни системное хранилище, ни запросы к
# другим хостам не затрагиваются, проверка TLS остаётся включённой — отключать её нельзя.
# Риск сознательный и узкий: данные публичные, без авторизации, а диапазонные гейты ниже
# ловят подменённые числа.
RS_ROOT_URL = "https://gu-st.ru/content/Other/doc/russian_trusted_root_ca.cer"
RS_SUB_URL = "http://nuc-cdp.voskhod.ru/cdp/subca_ssl_rsa2024.crt"
RS_PRICE_PAGE = "https://rosstat.gov.ru/statistics/price"
_RS_CTX = None


def _pem(raw: bytes) -> str:
    for form in ("DER", "PEM"):
        out = subprocess.run(["openssl", "x509", "-inform", form], input=raw,
                             capture_output=True, timeout=20, check=False).stdout
        text = out.decode("utf-8", "replace")
        if "BEGIN CERTIFICATE" in text:
            return text
    return ""


def rosstat_ctx():
    """SSL-контекст с корнем Минцифры. None, если собрать не удалось (тогда Росстат пропускаем)."""
    global _RS_CTX
    if _RS_CTX is not None:
        return _RS_CTX or None
    try:
        import certifi
        root = _pem(subprocess.run(["curl", "-fsS", "--max-time", "30", RS_ROOT_URL],
                                   capture_output=True, timeout=40, check=False).stdout)
        sub = _pem(subprocess.run(["curl", "-fsS", "--max-time", "30", RS_SUB_URL],
                                  capture_output=True, timeout=40, check=False).stdout)
        if not root:
            _RS_CTX = False
            return None
        fd, path = tempfile.mkstemp(prefix="rosstat-ca-", suffix=".pem")
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            fh.write(open(certifi.where(), encoding="utf-8").read())
            fh.write("\n" + root + ("\n" + sub if sub else ""))
        _RS_CTX = ssl.create_default_context(cafile=path)
        return _RS_CTX
    except Exception as e:  # noqa: BLE001
        sys.stderr.write(f"[macro] бандл для Росстата не собран: {e}\n")
        _RS_CTX = False
        return None


def rosstat_get(url: str, timeout: int = 60) -> bytes:
    ctx = rosstat_ctx()
    if ctx is None:
        raise RuntimeError("нет доверенного бандла для rosstat.gov.ru")
    req = urllib.request.Request(url, headers={**UA, "Accept-Language": "ru-RU,ru;q=0.9"})
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
        return r.read()


MONTHS_RU_ORDER = ["январь", "февраль", "март", "апрель", "май", "июнь",
                   "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
MOM_RANGE = (80.0, 130.0)      # индекс к предыдущему месяцу вне коридора — ошибка разбора
MIN_MOM_POINTS = 60


def fetch_monthly_mom() -> dict:
    """Официальный месячный ИПЦ Росстата «к концу предыдущего месяца» → прирост за месяц, %.

    Почему это отдельный источник: таблица ЦБ даёт инфляцию ТОЛЬКО год к году, а месячный
    прирост из неё математически не выводится без уровня индекса. Раньше мы честно писали,
    что «за месяц» показать нельзя; теперь берём официальный ряд Росстата.

    В файле лист «01» — строки-месяцы × колонки-годы, значения ИНДЕКСЫ (106.2 = +6,2%).
    Преобразование определяется типом поля источника (индекс к предыдущему месяцу), а не
    эвристикой «если больше 100»: вычитаем 100 ровно один раз.
    """
    html = rosstat_get(RS_PRICE_PAGE).decode("utf-8", "replace")
    m = re.findall(r'href="(/storage/mediabank/ipc_mes_[\d-]+\.xlsx)"', html)
    if not m:
        raise RuntimeError("на странице цен нет файла ipc_mes_*.xlsx")
    href = sorted(m)[-1]
    raw = rosstat_get("https://rosstat.gov.ru" + href, timeout=90)

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    if "01" not in wb.sheetnames:
        raise RuntimeError(f"лист «01» не найден: {wb.sheetnames[:5]}")
    ws = wb["01"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    years = None
    for row in rows[:8]:
        cand = [str(c) for c in row[1:] if isinstance(c, (int, float)) and 1990 <= int(c) <= 2100]
        if len(cand) > 10:
            years = row
            break
    if years is None:
        raise RuntimeError("строка с годами не найдена")

    out = []
    for row in rows:
        label = str(row[0] or "").strip().lower()
        if label not in MONTHS_RU_ORDER:
            continue
        mm = MONTHS_RU_ORDER.index(label) + 1
        for y, v in zip(years, row):
            if not isinstance(y, (int, float)) or not isinstance(v, (int, float)):
                continue
            year = int(y)
            if not (1990 <= year <= 2100) or not (MOM_RANGE[0] <= float(v) <= MOM_RANGE[1]):
                continue
            out.append({"month": f"{year}-{mm:02d}", "mom_pct": round(float(v) - 100.0, 4)})
    out.sort(key=lambda r: r["month"])
    if len(out) < MIN_MOM_POINTS:
        raise RuntimeError(f"месячный ИПЦ: {len(out)} точек, нужно ≥{MIN_MOM_POINTS}")
    return {"rows": out, "source_file": href, "latest": out[-1],
            "note": ("Официальный месячный ИПЦ Росстата: индекс к концу предыдущего месяца, "
                     "переведён в прирост (106,2 → +6,2%). Не интерполируется.")}


def fetch_key_rate(days: int = 900) -> list[dict]:
    """Дневная история ключевой ставки через SOAP KeyRate."""
    today = datetime.now(MSK).date()
    frm = today - timedelta(days=days)
    body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/"><soap:Body>'
        f'<KeyRate xmlns="{SOAP_NS}"><fromDate>{frm.isoformat()}</fromDate>'
        f'<ToDate>{today.isoformat()}</ToDate></KeyRate>'
        '</soap:Body></soap:Envelope>'
    ).encode("utf-8")
    xml = http(SOAP_URL, data=body, headers={
        "Content-Type": "text/xml; charset=utf-8", "SOAPAction": SOAP_NS + "KeyRate"}).decode("utf-8", "replace")
    rows = []
    for dt, rate in re.findall(r"<DT>([^<]+)</DT>\s*<Rate>([^<]+)</Rate>", xml):
        try:
            value = float(rate.replace(",", "."))
        except ValueError:
            continue
        if not (RATE_RANGE[0] <= value <= RATE_RANGE[1]):
            continue
        rows.append({"date": dt[:10], "rate": round(value, 4)})
    rows.sort(key=lambda r: r["date"])
    return rows


def fetch_inflation() -> list[dict]:
    """Месячная таблица ЦБ: ставка на конец месяца, инфляция г/г, цель."""
    today = datetime.now(MSK).date().strftime("%d.%m.%Y")
    url = f"{INFL_URL}?UniDbQuery.Posted=True&UniDbQuery.From={HISTORY_FROM}&UniDbQuery.To={today}"
    html = http(url, timeout=60).decode("utf-8", "replace")

    def num(text: str) -> float | None:
        cleaned = re.sub(r"[^\d,.\-]", "", text).replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None

    out = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S):
        cells = [re.sub(r"<[^>]+>|&nbsp;|\s+", " ", c).strip()
                 for c in re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S)]
        if len(cells) < 3 or not re.match(r"^\d{2}\.\d{4}$", cells[0]):
            continue
        mm, yyyy = cells[0].split(".")
        rate, infl = num(cells[1]), num(cells[2])
        target = num(cells[3]) if len(cells) > 3 else None
        if infl is None or not (INFL_RANGE[0] <= infl <= INFL_RANGE[1]):
            continue
        if rate is not None and not (RATE_RANGE[0] <= rate <= RATE_RANGE[1]):
            rate = None
        out.append({"month": f"{yyyy}-{mm}", "year": int(yyyy), "m": int(mm),
                    "key_rate": rate, "inflation_yoy": infl, "target": target})
    out.sort(key=lambda r: r["month"])
    return out


EXPECT_PAGE = "https://www.cbr.ru/statistics/ddkp/inflationary_expectations/"
EXPECT_SHEET = "Данные за все годы"
EXPECT_ROWS = {                     # подпись в первой колонке → ключ ряда
    "наблюдаемая инфляция (в %)": "perceived",
    "ожидаемая инфляция (в %)": "expected",
}
EXPECT_RANGE = (0.0, 60.0)          # медиана опроса вне коридора — ошибка разбора
MIN_EXPECT_POINTS = 60


def fetch_expectations() -> dict:
    """Инфляционные ожидания и наблюдаемая инфляция населением (опрос ФОМ по заказу ЦБ).

    Зачем инвестору: официальный ИПЦ и то, как люди ОЩУЩАЮТ рост цен, расходятся в разы
    (июль 2026: ИПЦ 6,0% против наблюдаемых 15,1%). ЦБ прямо ссылается на ожидания в
    решениях по ставке, поэтому разрыв «ожидания минус цель» объясняет жёсткость ДКП
    лучше, чем сам ИПЦ.

    Файл публикуется помесячно (Infl_exp_YY-MM.xlsx) и содержит ВСЮ историю, поэтому
    берём последний и не копим архив. Ссылку НЕ хардкодим — она меняется каждый месяц
    вместе с числовым id: ищем самую свежую на странице статистики.
    """
    html = http(EXPECT_PAGE, timeout=45).decode("utf-8", "replace")
    links = re.findall(r'href="(/Collection/Collection/File/\d+/Infl_exp_(\d\d)-(\d\d)\.xlsx)"', html)
    if not links:
        raise RuntimeError("на странице ожиданий нет ссылок Infl_exp_*.xlsx")
    href, yy, mm = max(links, key=lambda x: (x[1], x[2]))     # самый свежий период
    raw = http("https://www.cbr.ru" + href, timeout=90)

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    if EXPECT_SHEET not in wb.sheetnames:
        raise RuntimeError(f"лист «{EXPECT_SHEET}» не найден: {wb.sheetnames[:4]}")
    ws = wb[EXPECT_SHEET]

    dates, series = None, {}
    for row in ws.iter_rows(values_only=True):
        if dates is None:
            n = sum(1 for c in row if isinstance(c, datetime))
            if n > 20:
                dates = row
            continue
        label = str(row[0] or "").strip().lower()
        key = EXPECT_ROWS.get(label)
        # Метка «ожидаемая инфляция (в %)» встречается в файле ДВАЖДЫ: сначала годовая
        # («прямые оценки годовой инфляции»), ниже — ПЯТИЛЕТНЯЯ. Берём только первое
        # вхождение, иначе в блок «на 12 месяцев» уехал бы пятилетний ряд.
        if not key or key in series:
            continue
        points = []
        for d, v in zip(dates, row):
            if not isinstance(d, datetime) or not isinstance(v, (int, float)):
                continue
            if not (EXPECT_RANGE[0] <= float(v) <= EXPECT_RANGE[1]):
                continue
            points.append({"month": d.strftime("%Y-%m"), "value": round(float(v), 2)})
        points.sort(key=lambda x: x["month"])
        if points:
            series[key] = points
    wb.close()

    missing = [k for k in EXPECT_ROWS.values() if k not in series]
    if missing:
        raise RuntimeError(f"ряды не найдены: {missing}")
    for key, pts in series.items():
        if len(pts) < MIN_EXPECT_POINTS:
            raise RuntimeError(f"{key}: {len(pts)} точек, нужно ≥{MIN_EXPECT_POINTS}")
    return {
        "source_file": href, "period": f"20{yy}-{mm}",
        "expected": series["expected"], "perceived": series["perceived"],
        "latest_expected": series["expected"][-1], "latest_perceived": series["perceived"][-1],
        "note": ("Медианные оценки опроса населения (ФОМ по заказу Банка России): «ожидаемая» — "
                 "инфляция на 12 месяцев вперёд, «наблюдаемая» — за прошедшие 12 месяцев. Это "
                 "ВОСПРИЯТИЕ людей, а не измеренный индекс цен, и оно систематически выше ИПЦ."),
    }


def prev_key(key: str):
    """Прошлое значение из уже опубликованного файла: при сбое источника лучше показать
    вчерашние данные с честной датой, чем стереть блок. Ошибка одного ряда не уничтожает
    другой, поэтому last-known-good ведётся ПОКЛЮЧЕВО."""
    try:
        with open(OUT, encoding="utf-8") as fh:
            return (json.load(fh).get("inflation") or {}).get(key)
    except (OSError, ValueError):
        return None


def prev_expectations():
    return prev_key("expectations")


def main() -> int:
    generated = datetime.now(MSK).replace(microsecond=0)
    mom, mom_error = None, None
    try:
        # Росстат — отдельный источник и отдельный try: его сбой не должен обнулять
        # ставку, инфляцию г/г и ожидания от ЦБ.
        mom = fetch_monthly_mom()
    except Exception as e:  # noqa: BLE001
        mom_error = str(e)[:160]
        sys.stderr.write(f"[macro] месячный ИПЦ Росстата не получен: {mom_error}\n")
    expectations, expect_error = None, None
    try:
        # Отдельный try: ожидания — дополнительный ряд, их сбой не должен обнулять
        # ставку и инфляцию. Ошибка одного источника не уничтожает другой.
        expectations = fetch_expectations()
    except Exception as e:  # noqa: BLE001
        expect_error = str(e)[:160]
        sys.stderr.write(f"[macro] инфляционные ожидания не получены: {expect_error}\n")
    try:
        rates = fetch_key_rate()
        infl = fetch_inflation()
    except Exception as e:  # noqa: BLE001
        # last-good остаётся на месте: пустой/битый макро-блок хуже вчерашнего верного
        sys.stderr.write(f"[macro] источник ЦБ недоступен, {OUT} не перезаписан: {e}\n")
        return 1

    if len(rates) < MIN_RATE_ROWS:
        fail(f"ключевая ставка: получено {len(rates)} записей, нужно ≥{MIN_RATE_ROWS}")
    if len(infl) < MIN_INFL_ROWS:
        fail(f"инфляция: получено {len(infl)} месяцев, нужно ≥{MIN_INFL_ROWS}")

    current = rates[-1]
    # предыдущее ОТЛИЧНОЕ значение ставки — чтобы показать факт и дату решения ЦБ
    prev_rate, changed_on = None, None
    for row in reversed(rates[:-1]):
        if abs(row["rate"] - current["rate"]) > 1e-9:
            prev_rate = row["rate"]
            break
    for idx in range(len(rates) - 1, 0, -1):
        if abs(rates[idx]["rate"] - rates[idx - 1]["rate"]) > 1e-9:
            changed_on = rates[idx]["date"]
            break

    last_infl = infl[-1]
    real_rate = (round(current["rate"] - last_infl["inflation_yoy"], 2)
                 if last_infl["inflation_yoy"] is not None else None)

    # свод по годам: декабрьское значение = итог года (для 2026 — последний доступный месяц)
    by_year = {}
    for row in infl:
        by_year[row["year"]] = row
    year_summary = [{"year": y, "month": by_year[y]["month"],
                     "inflation_yoy": by_year[y]["inflation_yoy"],
                     "key_rate": by_year[y]["key_rate"],
                     "partial": by_year[y]["m"] != 12}
                    for y in sorted(by_year)]

    payload = {
        "schema_version": 1,
        "generated_at": generated.isoformat(),
        "source": {
            "key_rate": {"name": "Банк России, SOAP DailyInfoWebServ/KeyRate", "url": SOAP_URL},
            "inflation": {"name": "Банк России, «Инфляция и ключевая ставка»", "url": INFL_URL},
        },
        "key_rate": {
            "current": current["rate"],
            "asof": current["date"],
            "previous": prev_rate,
            "change": round(current["rate"] - prev_rate, 4) if prev_rate is not None else None,
            "changed_on": changed_on,
            "series": rates,
        },
        "inflation": {
            "latest_yoy": last_infl["inflation_yoy"],
            "latest_month": last_infl["month"],
            "target": last_infl["target"],
            "above_target": (round(last_infl["inflation_yoy"] - last_infl["target"], 2)
                             if last_infl["target"] is not None else None),
            "monthly": infl,
            "by_year": year_summary,
            # ожидания живут рядом с остальными рядами по ценам, а не отдельным
            # разделом: фронт читает их как inflation.expectations
            "expectations": expectations or prev_expectations(),
            "expectations_error": expect_error if expectations is None else None,
            "mom": mom or prev_key("mom"),
            "mom_error": mom_error if mom is None else None,
        },
        "real_key_rate": real_rate,
        "note": ("Ключевая ставка — дневная история Банка России. Инфляция — % год к году из "
                 "официальной таблицы ЦБ «Инфляция и ключевая ставка»; месячная инфляция ЦБ "
                 "публикуется с лагом, поэтому последний месяц может отставать от текущей даты. "
                 "Реальная ставка = ключевая ставка − инфляция г/г. Ничего не интерполируется."),
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    tmp = OUT + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, OUT)
    print(f"[macro] записано: site/macro_cbr.json · ставка {current['rate']}% на {current['date']}"
          f" (пред. {prev_rate}, решение {changed_on}) · инфляция {last_infl['inflation_yoy']}% "
          f"г/г за {last_infl['month']} · реальная {real_rate} п.п. · месяцев {len(infl)}, лет {len(year_summary)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
